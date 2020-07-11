import requests
import openpyxl as xl
from time import sleep


API_KEY = << << ENTER YOUR API KEY HERE >> >>


def setup_excel(workbook):
    """
    Setup sheets for the newly created excel file.
    """
    sheet1 = workbook.active
    sheet1.title = 'Sheet 1'
    sheet1['A1'] = 'City'
    sheet1['B1'] = 'Temperature'
    sheet1['C1'] = 'C/F'
    sheet1['D1'] = '0/1'
    sheet2 = workbook.create_sheet('Sheet 2')
    sheet2['A1'] = 'City'


def open_excel(file_name='weather.xlsx'):
    """
    Either opens an existing file or creates a new file and returns its
    workbook reference.
    """
    try:
        wb = xl.load_workbook(file_name)
    except FileNotFoundError:
        wb = xl.Workbook()
        setup_excel(wb)
        save_workbook(wb)
    return wb


def save_workbook(workbook, file_name='weather.xlsx'):
    """
    Saves workbook progress to excel file.
    """
    workbook.save(file_name)


def get_temperature(city):
    """
    Returns temperature in Celsius for a specified city.
    """
    url = 'https://api.openweathermap.org/data/2.5/weather'
    response = requests.get(url, params={'q': city, 'appid': API_KEY})
    stats = response.json()
    if 'main' in stats.keys():
        temperature = stats['main']['temp'] - 273.15
    else:
        temperature = 'NA'
    return temperature


def write_temperature(workbook, city, temp):
    """
    Writes updated value of temperature to workbook and excel file.
    """
    ws = workbook['Sheet 1']
    for row in ws.iter_rows():
        if row[0].value == city:
            row[1].value = temp
            break
    else:
        ws.append([city, temp, 'C', 1])
    save_workbook(workbook)


def get_cities(workbook):
    """
    It returns the list of rows from top row to the row above the blank row.
    """
    sheet1 = workbook['Sheet 1']
    return sheet1.iter_rows()


def celsius_to_fahrenheit(temperature):
    """
    Converts temperature value from celsius to fahrenheit.
    """
    return (temperature*9/5) + 32


def main():
    """
    As we get the list of cities we ping for that city's temperature if the
    value is 1.
    And we write the temperature of that city in Celsius or Fahrenheit as
    directed in excel.
    This task is repetitive unless we stop this program by force.
    """
    workbook = open_excel()
    while(True):
        for city_row in get_cities(workbook):
            if city_row[3].value == 1:
                temperature = get_temperature(city_row[0].value)
                if city_row[2].value == 'F':
                    temperature = celsius_to_fahrenheit(temperature)
                write_temperature(workbook, city_row[0].value, temperature)
        sleep(5)


if __name__ == '__main__':
    main()
